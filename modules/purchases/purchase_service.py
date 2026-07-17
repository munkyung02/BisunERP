import re
import sqlite3
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DATABASE_PATH = PROJECT_ROOT / "data" / "bisun_erp.db"
OUTPUT_ROOT = PROJECT_ROOT / "output" / "발주서"


class PurchaseService:
    """매핑 완료 주문을 공급처별 발주서로 생성합니다."""

    def __init__(
        self,
        database_path: str | Path | None = None,
        output_root: str | Path | None = None,
    ) -> None:
        self.database_path = Path(
            database_path or DATABASE_PATH
        )

        self.output_root = Path(
            output_root or OUTPUT_ROOT
        )

    # =========================================================
    # DB 연결
    # =========================================================

    def _connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(
            self.database_path
        )

        connection.row_factory = sqlite3.Row
        connection.execute(
            "PRAGMA foreign_keys = ON"
        )

        return connection

    # =========================================================
    # 발주 가능 상품 조회
    # =========================================================

    def get_purchase_candidates(
        self,
    ) -> list[dict[str, Any]]:
        """
        다음 조건을 만족하는 주문상품만 조회합니다.

        1. 상품 매핑 완료
        2. 공급처 연결 완료
        3. 아직 purchase_orders에 생성되지 않음
        4. 주문 발주 상태가 발주대기 또는 발주준비
        """

        query = """
            SELECT
                oi.id AS order_item_id,
                oi.order_id,
                oi.product_id,
                oi.supplier_id,
                oi.platform_product_name,
                oi.option_name,
                oi.quantity,
                oi.purchase_round,
                oi.mapping_status,

                o.platform,
                o.order_number,
                o.ordered_at,
                o.receiver_name,
                o.receiver_phone,
                o.postal_code,
                o.address,
                o.detail_address,
                o.delivery_message,
                o.purchase_status,

                p.product_code,
                p.product_name,
                p.supplier_product_name,
                p.purchase_price,

                s.supplier_code,
                s.supplier_name,
                s.contact_name,
                s.phone AS supplier_phone

            FROM order_items AS oi

            INNER JOIN orders AS o
                ON o.id = oi.order_id

            INNER JOIN products AS p
                ON p.id = oi.product_id

            INNER JOIN suppliers AS s
                ON s.id = oi.supplier_id

            LEFT JOIN purchase_orders AS po
                ON po.order_item_id = oi.id

            WHERE oi.product_id IS NOT NULL
              AND oi.supplier_id IS NOT NULL
              AND oi.mapping_status != '미매핑'
              AND o.purchase_status IN (
                    '발주대기',
                    '발주준비'
              )
              AND po.id IS NULL
              AND p.is_active = 1
              AND s.is_active = 1

            ORDER BY
                s.supplier_name,
                oi.purchase_round,
                o.ordered_at,
                o.id,
                oi.id
        """

        with self._connect() as connection:
            rows = connection.execute(
                query
            ).fetchall()

        return [
            dict(row)
            for row in rows
        ]

    # =========================================================
    # 발주 생성
    # =========================================================

    def create_purchase_files(
        self,
        order_item_ids: list[int] | None = None,
    ) -> dict[str, Any]:
        candidates = self.get_purchase_candidates()

        if order_item_ids is not None:
            selected_ids = {
                int(item_id)
                for item_id in order_item_ids
            }

            candidates = [
                candidate
                for candidate in candidates
                if int(candidate["order_item_id"])
                in selected_ids
            ]

        if order_item_ids is not None:
            selected_ids = {
                int(order_item_id)
                for order_item_id in order_item_ids
            }

            candidates = [
                candidate
                for candidate in candidates
                if int(candidate["order_item_id"])
                in selected_ids
            ]

        if not candidates:
            return {
                "candidate_count": 0,
                "created_count": 0,
                "supplier_count": 0,
                "files": [],
                "message": (
                    "발주 가능한 주문상품이 없습니다."
                ),
            }

        grouped: dict[
            tuple[int, str, str],
            list[dict[str, Any]],
        ] = defaultdict(list)

        for item in candidates:
            supplier_id = int(
                item["supplier_id"]
            )

            supplier_name = str(
                item["supplier_name"]
            ).strip()

            purchase_round = str(
                item.get("purchase_round") or "기본"
            ).strip()

            grouped[
                (
                    supplier_id,
                    supplier_name,
                    purchase_round,
                )
            ].append(item)

        date_folder = datetime.now().strftime(
            "%Y-%m-%d"
        )

        output_directory = (
            self.output_root / date_folder
        )

        output_directory.mkdir(
            parents=True,
            exist_ok=True,
        )

        created_files: list[str] = []
        created_order_item_ids: list[int] = []

        with self._connect() as connection:
            try:
                for (
                    supplier_id,
                    supplier_name,
                    purchase_round,
                ), items in grouped.items():
                    file_path = self._create_supplier_excel(
                        output_directory=output_directory,
                        supplier_name=supplier_name,
                        purchase_round=purchase_round,
                        items=items,
                    )

                    created_files.append(
                        str(file_path)
                    )

                    for item in items:
                        order_item_id = int(
                            item["order_item_id"]
                        )

                        full_address = self._full_address(
                            item.get("address"),
                            item.get(
                                "detail_address"
                            ),
                        )

                        connection.execute(
                            """
                            INSERT INTO purchase_orders (
                                supplier_id,
                                order_id,
                                order_item_id,
                                supplier_name,
                                order_number,
                                product_name,
                                option_name,
                                quantity,
                                receiver_name,
                                receiver_phone,
                                postal_code,
                                address,
                                delivery_message,
                                purchase_status,
                                purchase_file,
                                purchased_at,
                                created_at,
                                updated_at
                            )
                            VALUES (
                                ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                                ?, ?, ?,
                                '발주완료',
                                ?,
                                CURRENT_TIMESTAMP,
                                CURRENT_TIMESTAMP,
                                CURRENT_TIMESTAMP
                            )
                            """,
                            (
                                supplier_id,
                                item["order_id"],
                                order_item_id,
                                supplier_name,
                                item["order_number"],
                                (
                                    item.get(
                                        "supplier_product_name"
                                    )
                                    or item.get(
                                        "product_name"
                                    )
                                    or item.get(
                                        "platform_product_name"
                                    )
                                ),
                                item.get("option_name"),
                                item.get("quantity", 1),
                                item.get("receiver_name"),
                                item.get("receiver_phone"),
                                item.get("postal_code"),
                                full_address,
                                item.get(
                                    "delivery_message"
                                ),
                                str(file_path),
                            ),
                        )

                        created_order_item_ids.append(
                            order_item_id
                        )

                affected_order_ids = {
                    int(item["order_id"])
                    for item in candidates
                }

                for order_id in affected_order_ids:
                    remaining_row = connection.execute(
                        """
                        SELECT COUNT(*) AS remaining_count

                        FROM order_items AS oi

                        LEFT JOIN purchase_orders AS po
                            ON po.order_item_id = oi.id

                        WHERE oi.order_id = ?
                          AND po.id IS NULL
                        """,
                        (order_id,),
                    ).fetchone()

                    remaining_count = int(
                        remaining_row[
                            "remaining_count"
                        ]
                        or 0
                    )

                    next_status = (
                        "발주완료"
                        if remaining_count == 0
                        else "발주준비"
                    )

                    connection.execute(
                        """
                        UPDATE orders
                        SET
                            purchase_status = ?,
                            updated_at = CURRENT_TIMESTAMP
                        WHERE id = ?
                        """,
                        (
                            next_status,
                            order_id,
                        ),
                    )

                connection.commit()

            except Exception:
                connection.rollback()

                for created_file in created_files:
                    path = Path(created_file)

                    if path.exists():
                        try:
                            path.unlink()
                        except OSError:
                            pass

                raise

        return {
            "candidate_count": len(candidates),
            "created_count": len(
                created_order_item_ids
            ),
            "supplier_count": len(grouped),
            "files": created_files,
            "output_directory": str(
                output_directory
            ),
            "message": "발주서 생성 완료",
        }

    # =========================================================
    # 엑셀 생성
    # =========================================================

    def _create_supplier_excel(
        self,
        *,
        output_directory: Path,
        supplier_name: str,
        purchase_round: str,
        items: list[dict[str, Any]],
    ) -> Path:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "발주서"

        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A5"

        title = "비선상회 공급처 발주서"

        worksheet.merge_cells(
            "A1:K1"
        )

        title_cell = worksheet["A1"]
        title_cell.value = title
        title_cell.font = Font(
            size=18,
            bold=True,
        )
        title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.row_dimensions[1].height = 32

        worksheet["A2"] = "공급처"
        worksheet["B2"] = supplier_name
        worksheet["D2"] = "발주차수"
        worksheet["E2"] = purchase_round
        worksheet["G2"] = "생성일시"
        worksheet["H2"] = datetime.now().strftime(
            "%Y-%m-%d %H:%M:%S"
        )

        worksheet["A3"] = "총 주문상품"
        worksheet["B3"] = len(items)
        worksheet["D3"] = "총 수량"
        worksheet["E3"] = sum(
            int(item.get("quantity") or 0)
            for item in items
        )

        headers = [
            "번호",
            "주문번호",
            "공급처 상품명",
            "옵션",
            "수량",
            "수령인",
            "연락처",
            "우편번호",
            "주소",
            "배송메시지",
            "주문일시",
        ]

        header_row = 4

        thin_border = Border(
            left=Side(
                style="thin",
                color="B7B7B7",
            ),
            right=Side(
                style="thin",
                color="B7B7B7",
            ),
            top=Side(
                style="thin",
                color="B7B7B7",
            ),
            bottom=Side(
                style="thin",
                color="B7B7B7",
            ),
        )

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAD3",
        )

        for column_index, header in enumerate(
            headers,
            start=1,
        ):
            cell = worksheet.cell(
                row=header_row,
                column=column_index,
                value=header,
            )

            cell.font = Font(
                bold=True
            )

            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        for row_index, item in enumerate(
            items,
            start=1,
        ):
            excel_row = header_row + row_index

            product_name = (
                item.get(
                    "supplier_product_name"
                )
                or item.get("product_name")
                or item.get(
                    "platform_product_name"
                )
                or ""
            )

            full_address = self._full_address(
                item.get("address"),
                item.get("detail_address"),
            )

            values = [
                row_index,
                item.get("order_number"),
                product_name,
                item.get("option_name"),
                item.get("quantity"),
                item.get("receiver_name"),
                item.get("receiver_phone"),
                item.get("postal_code"),
                full_address,
                item.get("delivery_message"),
                item.get("ordered_at"),
            ]

            for column_index, value in enumerate(
                values,
                start=1,
            ):
                cell = worksheet.cell(
                    row=excel_row,
                    column=column_index,
                    value=value or "",
                )

                cell.border = thin_border
                cell.alignment = Alignment(
                    horizontal=(
                        "center"
                        if column_index in {
                            1,
                            2,
                            5,
                            6,
                            7,
                            8,
                            11,
                        }
                        else "left"
                    ),
                    vertical="center",
                    wrap_text=True,
                )

            worksheet.row_dimensions[
                excel_row
            ].height = 38

        column_widths = {
            1: 7,
            2: 18,
            3: 28,
            4: 28,
            5: 8,
            6: 11,
            7: 16,
            8: 10,
            9: 48,
            10: 38,
            11: 19,
        }

        for column_index, width in (
            column_widths.items()
        ):
            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = width

        worksheet.auto_filter.ref = (
            f"A4:K{header_row + len(items)}"
        )

        safe_supplier_name = (
            self._safe_filename(
                supplier_name
            )
        )

        safe_purchase_round = (
            self._safe_filename(
                purchase_round
            )
        )

        timestamp = datetime.now().strftime(
            "%H%M%S"
        )

        file_name = (
            f"{safe_supplier_name}_"
            f"{safe_purchase_round}_"
            f"발주서_{timestamp}.xlsx"
        )

        file_path = (
            output_directory / file_name
        )

        workbook.save(file_path)

        return file_path

    # =========================================================
    # 조회용
    # =========================================================

    def get_purchase_orders(
        self,
    ) -> list[dict[str, Any]]:
        with self._connect() as connection:
            rows = connection.execute(
                """
                SELECT
                    po.*,
                    o.platform,
                    o.ordered_at,
                    s.supplier_code

                FROM purchase_orders AS po

                INNER JOIN orders AS o
                    ON o.id = po.order_id

                LEFT JOIN suppliers AS s
                    ON s.id = po.supplier_id

                ORDER BY
                    po.created_at DESC,
                    po.id DESC
                """
            ).fetchall()

        return [
            dict(row)
            for row in rows
        ]

    # =========================================================
    # 공통 함수
    # =========================================================

    @staticmethod
    def _full_address(
        address: Any,
        detail_address: Any,
    ) -> str:
        values = [
            str(value).strip()
            for value in (
                address,
                detail_address,
            )
            if value not in (None, "")
        ]

        return " ".join(values)

    @staticmethod
    def _safe_filename(
        value: str,
    ) -> str:
        cleaned = re.sub(
            r'[\\/:*?"<>|]',
            "_",
            str(value).strip(),
        )

        cleaned = re.sub(
            r"\s+",
            "_",
            cleaned,
        )

        return cleaned or "미지정"