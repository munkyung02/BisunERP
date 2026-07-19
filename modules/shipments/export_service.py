from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from modules.shipments.shipment_repository import (
    ShipmentRepository,
)


class ExportService:
    """쿠팡 송장등록용 엑셀파일 생성 서비스입니다."""

    COUPANG_SHEET_NAME = "Delivery"
    COUPANG_HEADER_ROW = 1
    COUPANG_DATA_START_ROW = 2

    REQUIRED_HEADERS = {
        "order_number": "주문번호",
        "courier_name": "택배사",
        "tracking_number": "운송장번호",
    }

    COURIER_NAMES = {
        "CJ대한통운": "CJ 대한통운",
        "대한통운": "CJ 대한통운",
        "CJ": "CJ 대한통운",
        "CJ택배": "CJ 대한통운",
        "CJ 대한통운": "CJ 대한통운",

        "한진": "한진택배",
        "한진택배": "한진택배",

        "롯데": "롯데택배",
        "롯데택배": "롯데택배",
        "롯데글로벌로지스": "롯데택배",

        "로젠": "로젠택배",
        "로젠택배": "로젠택배",

        "우체국": "우체국택배",
        "우체국택배": "우체국택배",

        "경동": "경동택배",
        "경동택배": "경동택배",

        "대신": "대신택배",
        "대신택배": "대신택배",

        "천일": "천일택배",
        "천일택배": "천일택배",

        "일양": "일양로지스",
        "일양택배": "일양로지스",
        "일양로지스": "일양로지스",
    }

    def __init__(self) -> None:
        self.repository = ShipmentRepository()

    def export_coupang_delivery_file(
        self,
        source_file_path: str | Path,
        output_directory: str | Path | None = None,
    ) -> dict[str, Any]:
        """
        쿠팡에서 내려받은 Delivery 파일에 ERP 송장을 입력하고
        새로운 송장등록용 엑셀파일을 생성합니다.
        """

        source_path = Path(
            source_file_path
        )

        if not source_path.exists():
            raise FileNotFoundError(
                "쿠팡 엑셀파일을 찾을 수 없습니다.\n"
                f"{source_path}"
            )

        if source_path.suffix.lower() not in {
            ".xlsx",
            ".xlsm",
        }:
            raise ValueError(
                "쿠팡 파일은 .xlsx 또는 .xlsm 형식이어야 합니다."
            )

        shipments = (
            self.repository
            .get_ready_shipments()
        )

        if not shipments:
            raise ValueError(
                "쿠팡 파일에 입력할 배송중 송장이 없습니다.\n"
                "먼저 공급처 송장을 분석하고 저장해 주세요."
            )

        shipment_map = (
            self._build_shipment_map(
                shipments
            )
        )

        keep_vba = (
            source_path.suffix.lower()
            == ".xlsm"
        )

        workbook = load_workbook(
            filename=source_path,
            keep_vba=keep_vba,
        )

        try:
            worksheet = self._get_delivery_sheet(
                workbook
            )

            header_map = self._build_header_map(
                worksheet
            )

            self._validate_headers(
                header_map
            )

            order_column = header_map[
                self.REQUIRED_HEADERS[
                    "order_number"
                ]
            ]

            courier_column = header_map[
                self.REQUIRED_HEADERS[
                    "courier_name"
                ]
            ]

            tracking_column = header_map[
                self.REQUIRED_HEADERS[
                    "tracking_number"
                ]
            ]

            total_order_count = 0
            matched_count = 0
            unmatched_count = 0
            already_registered_count = 0

            unmatched_orders: list[str] = []
            matched_orders: list[str] = []

            for excel_row in range(
                self.COUPANG_DATA_START_ROW,
                worksheet.max_row + 1,
            ):
                order_number = self._clean_text(
                    worksheet.cell(
                        row=excel_row,
                        column=order_column,
                    ).value
                )

                if not order_number:
                    continue

                total_order_count += 1

                current_tracking_number = (
                    self._clean_text(
                        worksheet.cell(
                            row=excel_row,
                            column=tracking_column,
                        ).value
                    )
                )

                if current_tracking_number:
                    already_registered_count += 1
                    continue

                shipment = shipment_map.get(
                    order_number
                )

                if shipment is None:
                    unmatched_count += 1
                    unmatched_orders.append(
                        order_number
                    )
                    continue

                courier_name = (
                    self._convert_courier_name(
                        shipment.get(
                            "courier_name",
                            "",
                        )
                    )
                )

                tracking_number = (
                    self._clean_tracking_number(
                        shipment.get(
                            "tracking_number",
                            "",
                        )
                    )
                )

                if not courier_name:
                    unmatched_count += 1
                    unmatched_orders.append(
                        order_number
                    )
                    continue

                if not tracking_number:
                    unmatched_count += 1
                    unmatched_orders.append(
                        order_number
                    )
                    continue

                worksheet.cell(
                    row=excel_row,
                    column=courier_column,
                ).value = courier_name

                worksheet.cell(
                    row=excel_row,
                    column=tracking_column,
                ).value = tracking_number

                matched_count += 1
                matched_orders.append(
                    order_number
                )

            output_path = (
                self._build_output_path(
                    source_path=source_path,
                    output_directory=output_directory,
                    suffix=source_path.suffix,
                )
            )

            workbook.save(
                output_path
            )

        finally:
            workbook.close()

        return {
            "source_file_path": str(
                source_path
            ),
            "output_file_path": str(
                output_path
            ),
            "shipment_count": len(
                shipments
            ),
            "total_order_count": (
                total_order_count
            ),
            "matched_count": matched_count,
            "unmatched_count": (
                unmatched_count
            ),
            "already_registered_count": (
                already_registered_count
            ),
            "matched_orders": (
                matched_orders
            ),
            "unmatched_orders": (
                unmatched_orders
            ),
        }

    def _build_shipment_map(
        self,
        shipments: list[dict[str, Any]],
    ) -> dict[str, dict[str, Any]]:
        """
        ERP 송장을 주문번호를 키로 하는 사전으로 변환합니다.

        하나의 주문번호에 여러 상품 송장이 존재하면
        송장번호를 슬래시로 합쳐서 처리합니다.
        """

        shipment_map: dict[
            str,
            dict[str, Any],
        ] = {}

        for shipment in shipments:
            order_number = self._clean_text(
                shipment.get(
                    "order_number"
                )
            )

            if not order_number:
                continue

            courier_name = self._clean_text(
                shipment.get(
                    "courier_name"
                )
            )

            tracking_number = (
                self._clean_tracking_number(
                    shipment.get(
                        "tracking_number"
                    )
                )
            )

            if not tracking_number:
                continue

            existing = shipment_map.get(
                order_number
            )

            if existing is None:
                shipment_map[
                    order_number
                ] = {
                    "order_number": (
                        order_number
                    ),
                    "courier_name": (
                        courier_name
                    ),
                    "tracking_number": (
                        tracking_number
                    ),
                }
                continue

            existing_numbers = (
                self._split_tracking_numbers(
                    existing[
                        "tracking_number"
                    ]
                )
            )

            new_numbers = (
                self._split_tracking_numbers(
                    tracking_number
                )
            )

            for number in new_numbers:
                if number not in existing_numbers:
                    existing_numbers.append(
                        number
                    )

            existing[
                "tracking_number"
            ] = "/".join(
                existing_numbers
            )

            if (
                not existing.get(
                    "courier_name"
                )
                and courier_name
            ):
                existing[
                    "courier_name"
                ] = courier_name

        return shipment_map

    def _get_delivery_sheet(
        self,
        workbook: Any,
    ) -> Any:
        if (
            self.COUPANG_SHEET_NAME
            in workbook.sheetnames
        ):
            return workbook[
                self.COUPANG_SHEET_NAME
            ]

        if not workbook.sheetnames:
            raise ValueError(
                "쿠팡 엑셀파일에 시트가 없습니다."
            )

        return workbook[
            workbook.sheetnames[0]
        ]

    def _build_header_map(
        self,
        worksheet: Any,
    ) -> dict[str, int]:
        header_map: dict[str, int] = {}

        for column_number in range(
            1,
            worksheet.max_column + 1,
        ):
            header = self._clean_text(
                worksheet.cell(
                    row=self.COUPANG_HEADER_ROW,
                    column=column_number,
                ).value
            )

            if header:
                header_map[
                    header
                ] = column_number

        return header_map

    def _validate_headers(
        self,
        header_map: dict[str, int],
    ) -> None:
        missing_headers: list[str] = []

        for header_name in (
            self.REQUIRED_HEADERS.values()
        ):
            if header_name not in header_map:
                missing_headers.append(
                    header_name
                )

        if missing_headers:
            missing_text = ", ".join(
                missing_headers
            )

            raise ValueError(
                "쿠팡 파일에서 필수 열을 찾지 못했습니다.\n"
                f"누락된 열: {missing_text}"
            )

    def _convert_courier_name(
        self,
        courier_name: Any,
    ) -> str:
        original_name = self._clean_text(
            courier_name
        )

        normalized_name = (
            original_name.replace(
                " ",
                "",
            )
        )

        if original_name in self.COURIER_NAMES:
            return self.COURIER_NAMES[
                original_name
            ]

        if normalized_name in self.COURIER_NAMES:
            return self.COURIER_NAMES[
                normalized_name
            ]

        return original_name

    def _clean_tracking_number(
        self,
        tracking_number: Any,
    ) -> str:
        text = self._clean_text(
            tracking_number
        )

        if text.endswith(".0"):
            text = text[:-2]

        text = text.replace(
            " ",
            "",
        )

        return text

    def _split_tracking_numbers(
        self,
        tracking_number: Any,
    ) -> list[str]:
        text = self._clean_tracking_number(
            tracking_number
        )

        if not text:
            return []

        numbers: list[str] = []

        normalized_text = (
            text.replace(
                ",",
                "/",
            )
            .replace(
                ";",
                "/",
            )
            .replace(
                "\n",
                "/",
            )
        )

        for part in normalized_text.split(
            "/"
        ):
            number = part.strip()

            if number and number not in numbers:
                numbers.append(
                    number
                )

        return numbers

    def _build_output_path(
        self,
        *,
        source_path: Path,
        output_directory: str | Path | None,
        suffix: str,
    ) -> Path:
        if output_directory is None:
            destination_directory = (
                source_path.parent
            )
        else:
            destination_directory = Path(
                output_directory
            )

        destination_directory.mkdir(
            parents=True,
            exist_ok=True,
        )

        current_time = datetime.now()

        date_text = current_time.strftime(
            "%Y%m%d"
        )
        time_text = current_time.strftime(
            "%H%M%S"
        )

        output_name = (
            f"쿠팡_송장등록_"
            f"{date_text}_"
            f"{time_text}"
            f"{suffix}"
        )

        return (
            destination_directory
            / output_name
        )

    def _clean_text(
        self,
        value: Any,
    ) -> str:
        if value is None:
            return ""

        if (
            isinstance(value, float)
            and value.is_integer()
        ):
            return str(
                int(value)
            ).strip()

        return str(
            value
        ).strip()