from __future__ import annotations

import re
from modules.shipments.shipment_repository import ShipmentRepository
from modules.orders.order_repository import OrderRepository
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


class ShipmentService:
    """공급처 송장 회신 엑셀 처리 서비스."""

    def __init__(self):

        self.shipment_repository = ShipmentRepository()
        self.order_repository = OrderRepository()

    SUPPLIER_PROFILES: dict[str, dict[str, Any]] = {
        "해담": {
            "sheet_name": "Sheet1",
            "header_row": 1,
            "columns": {
                "receiver_name": "받는분성명",
                "receiver_phone": "받는분전화번호",
                "receiver_address": "받는분주소(전체, 분할)",
                "product_name": "품목",
                "quantity": "수량",
                "delivery_message": "배송메세지",

                # 해담 파일은 실제 데이터가 제목과 한 칸씩 어긋나 있음
                "carrier_column": 10,          # J열
                "tracking_number_column": 11,  # K열
            },
        },
    }

    def match_shipment(
        self,
        shipment: dict[str, Any],
    ) -> dict[str, Any]:

        candidates = self.shipment_repository.find_match_candidates(
            supplier_name=shipment["supplier_name"],
            receiver_name=shipment["receiver_name"],
            receiver_phone=shipment["receiver_phone"],
            quantity=shipment["quantity"],
        )

        if not candidates:
            return {
                "status": "unmatched",
                "shipment": shipment,
                "candidates": [],
            }

        if len(candidates) > 1:
            return {
                "status": "ambiguous",
                "shipment": shipment,
                "candidates": candidates,
            }

        return {
            "status": "matched",
            "shipment": shipment,
            "candidate": candidates[0],
            "candidates": candidates,
        }
    
    def preview_supplier_shipment_file(
        self,
        file_path: str | Path,
        supplier_name: str,
    ) -> dict[str, Any]:
        """송장파일을 읽고 발주내역과 자동매칭합니다."""

        file_result = (
            self.read_supplier_shipment_file(
                file_path=file_path,
                supplier_name=supplier_name,
            )
        )

        match_results = []

        matched_count = 0
        unmatched_count = 0
        ambiguous_count = 0
        duplicate_count = 0

        for shipment in file_result["rows"]:
            tracking_number = shipment[
                "tracking_number"
            ]

            if self.shipment_repository.shipment_exists(
                tracking_number
            ):
                match_result = {
                    "status": "duplicate",
                    "shipment": shipment,
                    "candidates": [],
                }

                duplicate_count += 1

            else:
                match_result = self.match_shipment(
                    shipment
                )

                if match_result["status"] == "matched":
                    matched_count += 1

                elif match_result["status"] == "ambiguous":
                    ambiguous_count += 1

                else:
                    unmatched_count += 1

            match_results.append(
                match_result
            )

        return {
            **file_result,
            "match_results": match_results,
            "matched_count": matched_count,
            "unmatched_count": unmatched_count,
            "ambiguous_count": ambiguous_count,
            "duplicate_count": duplicate_count,
        }

    def save_matched_shipments(
        self,
        match_results: list[dict[str, Any]],
    ) -> dict[str, Any]:
        """자동매칭에 성공한 송장을 DB에 저장합니다."""

        saved_count = 0
        skipped_count = 0
        error_count = 0
        errors = []

        for result in match_results:
            if result.get("status") != "matched":
                skipped_count += 1
                continue

            shipment = result["shipment"]
            candidate = result["candidate"]

            tracking_number = shipment[
                "tracking_number"
            ]

            try:
                if (
                    self.shipment_repository
                    .shipment_exists(
                        tracking_number
                    )
                ):
                    skipped_count += 1
                    continue

                self.shipment_repository.create_shipment(
                    order_id=int(
                        candidate["order_id"]
                    ),
                    order_item_id=int(
                        candidate["order_item_id"]
                    ),
                    supplier_id=(
                        int(candidate["supplier_id"])
                        if candidate.get("supplier_id")
                        is not None
                        else None
                    ),
                    courier_name=shipment[
                        "carrier"
                    ],
                    tracking_number=tracking_number,
                    shipment_status="배송중",
                )

                self.shipment_repository.update_purchase_order_status(
                    purchase_order_id=int(
                        candidate[
                            "purchase_order_id"
                        ]
                    ),
                    purchase_status="배송중",
                )

                saved_count += 1

            except Exception as error:
                error_count += 1

                errors.append(
                    {
                        "excel_row": shipment.get(
                            "excel_row"
                        ),
                        "tracking_number": (
                            tracking_number
                        ),
                        "error": str(error),
                    }
                )

        return {
            "saved_count": saved_count,
            "skipped_count": skipped_count,
            "error_count": error_count,
            "errors": errors,
        }

    def read_supplier_shipment_file(
        self,
        file_path: str | Path,
        supplier_name: str,
    ) -> dict[str, Any]:
        """
        공급처가 회신한 엑셀에서 송장 정보를 읽는다.

        반환 형식:
        {
            "supplier_name": "해담",
            "file_path": "...",
            "rows": [...],
            "errors": [...],
            "total_count": 10,
            "valid_count": 9,
            "error_count": 1,
        }
        """
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(
                f"엑셀 파일을 찾을 수 없습니다.\n{path}"
            )

        profile = self.SUPPLIER_PROFILES.get(supplier_name)

        if profile is None:
            raise ValueError(
                f"등록되지 않은 공급처 양식입니다: {supplier_name}"
            )

        workbook = load_workbook(
            filename=path,
            data_only=True,
            read_only=True,
        )

        try:
            worksheet = self._get_worksheet(
                workbook=workbook,
                sheet_name=profile["sheet_name"],
            )

            header_row = int(profile["header_row"])
            header_map = self._build_header_map(
                worksheet=worksheet,
                header_row=header_row,
            )

            rows: list[dict[str, Any]] = []
            errors: list[dict[str, Any]] = []

            for excel_row in range(
                header_row + 1,
                worksheet.max_row + 1,
            ):
                parsed_row = self._parse_haedam_row(
                    worksheet=worksheet,
                    excel_row=excel_row,
                    header_map=header_map,
                    profile=profile,
                )

                if parsed_row is None:
                    continue

                validation_errors = self._validate_shipment_row(
                    parsed_row
                )

                if validation_errors:
                    parsed_row["status"] = "오류"
                    parsed_row["error_message"] = ", ".join(
                        validation_errors
                    )
                    errors.append(parsed_row)
                else:
                    parsed_row["status"] = "정상"
                    parsed_row["error_message"] = ""
                    rows.append(parsed_row)

            return {
                "supplier_name": supplier_name,
                "file_path": str(path),
                "rows": rows,
                "errors": errors,
                "total_count": len(rows) + len(errors),
                "valid_count": len(rows),
                "error_count": len(errors),
            }

        finally:
            workbook.close()

    def _get_worksheet(
        self,
        workbook: Any,
        sheet_name: str,
    ) -> Any:
        if sheet_name in workbook.sheetnames:
            return workbook[sheet_name]

        return workbook[workbook.sheetnames[0]]

    def _build_header_map(
        self,
        worksheet: Any,
        header_row: int,
    ) -> dict[str, int]:
        header_map: dict[str, int] = {}

        for column_number in range(
            1,
            worksheet.max_column + 1,
        ):
            value = worksheet.cell(
                row=header_row,
                column=column_number,
            ).value

            header = self._clean_text(value)

            if header:
                header_map[header] = column_number

        return header_map

    def _parse_haedam_row(
        self,
        worksheet: Any,
        excel_row: int,
        header_map: dict[str, int],
        profile: dict[str, Any],
    ) -> dict[str, Any] | None:
        columns = profile["columns"]

        receiver_name = self._get_value_by_header(
            worksheet,
            excel_row,
            header_map,
            columns["receiver_name"],
        )
        receiver_phone = self._get_value_by_header(
            worksheet,
            excel_row,
            header_map,
            columns["receiver_phone"],
        )
        receiver_address = self._get_value_by_header(
            worksheet,
            excel_row,
            header_map,
            columns["receiver_address"],
        )
        product_name = self._get_value_by_header(
            worksheet,
            excel_row,
            header_map,
            columns["product_name"],
        )
        quantity = self._get_value_by_header(
            worksheet,
            excel_row,
            header_map,
            columns["quantity"],
        )
        delivery_message = self._get_value_by_header(
            worksheet,
            excel_row,
            header_map,
            columns["delivery_message"],
        )

        carrier = worksheet.cell(
            row=excel_row,
            column=columns["carrier_column"],
        ).value

        tracking_value = worksheet.cell(
            row=excel_row,
            column=columns["tracking_number_column"],
        ).value

        # 완전히 비어 있는 행은 무시
        important_values = [
            receiver_name,
            receiver_phone,
            product_name,
            carrier,
            tracking_value,
        ]

        if all(
            self._clean_text(value) == ""
            for value in important_values
        ):
            return None

        tracking_numbers = self._split_tracking_numbers(
            tracking_value
        )

        return {
            "excel_row": excel_row,
            "supplier_name": "해담",
            "receiver_name": self._clean_text(
                receiver_name
            ),
            "receiver_phone": self._normalize_phone(
                receiver_phone
            ),
            "receiver_address": self._clean_text(
                receiver_address
            ),
            "product_name": self._clean_text(
                product_name
            ),
            "quantity": self._normalize_quantity(
                quantity
            ),
            "delivery_message": self._clean_text(
                delivery_message
            ),
            "carrier": self._normalize_carrier(
                carrier
            ),
            "tracking_numbers": tracking_numbers,
            "tracking_number": "/".join(tracking_numbers),
        }

    def _get_value_by_header(
        self,
        worksheet: Any,
        excel_row: int,
        header_map: dict[str, int],
        header_name: str,
    ) -> Any:
        column_number = header_map.get(header_name)

        if column_number is None:
            return None

        return worksheet.cell(
            row=excel_row,
            column=column_number,
        ).value

    def _validate_shipment_row(
        self,
        row: dict[str, Any],
    ) -> list[str]:
        errors: list[str] = []

        if not row["receiver_name"]:
            errors.append("수취인명 없음")

        if not row["receiver_phone"]:
            errors.append("전화번호 없음")

        if not row["product_name"]:
            errors.append("상품명 없음")

        if row["quantity"] <= 0:
            errors.append("수량 오류")

        if not row["carrier"]:
            errors.append("택배사 없음")

        if not row["tracking_numbers"]:
            errors.append("송장번호 없음")

        return errors

    def _split_tracking_numbers(
        self,
        value: Any,
    ) -> list[str]:
        text = self._clean_text(value)

        if not text:
            return []

        # 엑셀에서 숫자로 저장된 경우 끝에 붙는 .0 제거
        if text.endswith(".0"):
            text = text[:-2]

        parts = re.split(
            r"[/,\n;]+",
            text,
        )

        tracking_numbers: list[str] = []

        for part in parts:
            number = re.sub(
                r"\s+",
                "",
                part,
            )

            if number and number not in tracking_numbers:
                tracking_numbers.append(number)

        return tracking_numbers

    def _normalize_carrier(
        self,
        value: Any,
    ) -> str:
        carrier = self._clean_text(value).replace(" ", "")

        carrier_aliases = {
            "대한통운": "CJ대한통운",
            "CJ": "CJ대한통운",
            "CJ택배": "CJ대한통운",
            "CJ대한통운": "CJ대한통운",
            "로젠": "로젠택배",
            "로젠택배": "로젠택배",
            "한진": "한진택배",
            "한진택배": "한진택배",
            "롯데": "롯데택배",
            "롯데택배": "롯데택배",
            "롯데글로벌로지스": "롯데택배",
            "우체국": "우체국택배",
            "우체국택배": "우체국택배",
        }

        return carrier_aliases.get(
            carrier,
            self._clean_text(value),
        )

    def _normalize_phone(
        self,
        value: Any,
    ) -> str:
        phone = self._clean_text(value)

        return re.sub(
            r"[^0-9]",
            "",
            phone,
        )

    def _normalize_quantity(
        self,
        value: Any,
    ) -> int:
        if value is None:
            return 0

        try:
            return int(float(value))
        except (TypeError, ValueError):
            return 0

    def _clean_text(
        self,
        value: Any,
    ) -> str:
        if value is None:
            return ""

        if isinstance(value, float) and value.is_integer():
            return str(int(value)).strip()

        return str(value).strip()